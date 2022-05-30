import pandas as pd
import openpyxl
import os
import sys
from pathlib import Path
import math
from scipy.optimize import curve_fit
import numpy as np
from matplotlib import pyplot as plt
import random
import seaborn as sns
from datetime import datetime
import csv


# get data from table with multipliers
cutoff_multiplier_accuracies = [95.0, 97.5, 99.0, 99.5, 99.9]
sdm = Path('Standard deviation multipliers.xlsx')
table_obj = openpyxl.load_workbook(sdm)
sheet = table_obj.active
rows = []
for row in sheet.iter_rows(2, sheet.max_row):
    rows.append([x.value for x in row[1:]])
multipliers = pd.DataFrame(rows, index=[x for x in range(2, 21)] + [25, 30], columns=cutoff_multiplier_accuracies)

markers = ['^', '3', 'P', 'x', 'v', 'p', '+', 'o', 's', 'd', '>', 'D']
colors = ['violet', 'purple', 'blue', 'lime', 'aqua', 'gold', 'chartreuse',
          'orangered', 'firebrick', 'dodgerblue', 'indigo', 'darkorange']


# один столбик с данными, которые относятся к одному образцу
class Sample:
    def __init__(self, name: str, xdata: list, ydata: list):
        self.name = name
        self.xdata = xdata
        self.ydata = ydata
        self.popt = None
        self.pcov = None
        self.endpoint_titer = None
        self.R2 = None
        self.bad_data = False
        self.plate = None
        self.group = None

        # if ydata is ascending, invert it
        if self.ydata[0] < self.ydata[-1]:
            self.ydata = self.ydata[::-1]

    def reverse_sigmoid(self, x, a, b):
        assert a >= 0.0, f'{a=} while should be non-negative'
        return min(self.ydata) + (max(self.ydata) - min(self.ydata))/(1+10**((math.log(a) - x) * b))

    def get_popt_pcov(self):
        self.popt, self.pcov = curve_fit(self.reverse_sigmoid, self.xdata, self.ydata, method='dogbox')

    def calculate_endpoint_titer(self, cutoff: float):
        self.get_popt_pcov()

        if max(self.ydata) > cutoff*2:
            self.endpoint_titer = 1/10**self.revert_x(cutoff*2)
        else:
            self.bad_data = True
            print(f'Sample {self.name} has a bad data for this cutoff ({cutoff}) and will not be included'
                  f' in calculations! Try to lower calculation accuracy.')

    def approximate(self, x: float) -> float:
        return self.reverse_sigmoid(x, self.popt[0], self.popt[1])

    def revert_x(self, y: float) -> float:
        a = self.popt[0]
        b = self.popt[1]
        return (math.log10((max(self.ydata) - y) / y) / b) - math.log(a)

    def get_endpoint_titer(self):
        return self.endpoint_titer

    def get_R2(self):
        residuals = []
        for i in range(0, len(self.ydata)):
            residuals.append(self.ydata[i] - self.reverse_sigmoid(self.xdata[i], self.popt[0], self.popt[1]))
        ss_res = np.sum([x**2 for x in residuals])
        ss_tot = np.sum([(y-np.mean(self.ydata))**2 for y in self.ydata])
        self.R2 = 1 - (ss_res/ss_tot)

    def get_quality_vector(self):
        return math.sqrt(self.approximate(math.log(self.popt[0]))**2 + math.log(self.popt[0])**2 + self.R2**2)

    def __repr__(self):
        return f'Sample "{self.name}"\n' + \
               f'xdata: {self.xdata}\n' + \
               f'ydata: {self.ydata}\n' + \
               f'popt: {self.popt}\npcov: {self.pcov}\n' + \
               f'endpoint_titer: {self.endpoint_titer}'


# группа образцов
class AnalyticalGroup:
    def __init__(self, name: str, samples: list):
        self.name = name
        self.samples = samples
        self.cutoff = None
        self.average_titer = None
        self.outliers = list()
        self.negative_control_indices = list()

    def add_sample(self, sample):
        self.samples.append(sample)

    def get_sample_by_name(self, name: str):
        for sample in self.samples:
            if sample.name == name:
                return sample

    # это должны быть ячейки, которые выберет юзер в качестве контроля для каждой группы,
    # но по умолчанию - все, что идет последним в образце
    def get_group_cutoff(self, accuracy: float):
        min_ydata = []
        for sample in self.samples:
            for index in self.negative_control_indices:
                min_ydata.append(sample.ydata[index])

        if len(min_ydata) > 1:
            stdev = np.std(min_ydata)
            mean = np.mean(min_ydata)
            # выбор тончости уровня значимости (1-a) нужно будет дать юзеру возможность выбирать (бегунком, например)
            t = multipliers[accuracy][len(min_ydata)]
            self.cutoff = mean + stdev*t
        else:
            self.cutoff = min_ydata[0]

        for sample in self.samples:
            sample.calculate_endpoint_titer(self.cutoff)

    def calculate_average_titer(self):
        titers = []
        for sample in self.samples:
            endpoint_titer = sample.get_endpoint_titer()
            if endpoint_titer is not None:
                titers.append(endpoint_titer)
        self.average_titer = sum(titers)/len(titers)

    def detect_outliers(self):
        group_vectors = {}
        for sample in self.samples:
            group_vectors[sample.name] = sample.get_quality_vector()

        outliers = detect_outlier(group_vectors.values(), 30, 70)
        keys = []
        for k, v in group_vectors.items():
            for outlier in outliers:
                if v == outlier:
                    keys.append(k)
        self.outliers = keys

    def plot_samples_data(self, folder_name=None, with_outliers=False):
        fig, ax = plt.subplots()

        colors_set = random.sample(colors, len(self.samples))
        markers_set = random.sample(markers, len(self.samples))
        index = 0
        samples = self.samples
        if with_outliers:
            for outlier in self.outliers:
                samples.remove(self.get_sample_by_name(outlier))
        for sample in samples:
            point_count = 80
            x_step = (sample.xdata[-1] - sample.xdata[0]) / float(point_count)
            interpolated_x = []
            for step_index in range(point_count + 1):
                interpolated_x.append(sample.xdata[0] + x_step * step_index)
            approximate_y = [sample.approximate(x) for x in interpolated_x]

            plt.scatter(sample.xdata, sample.ydata, marker=markers_set[index],
                        color=colors_set[index], label=f'{sample.name} R^2={sample.R2:.3f}')
            plt.plot(interpolated_x, approximate_y, linestyle='-', color=colors_set[index])
            index += 1

        ax.spines['left'].set_color('darkblue')
        ax.spines['bottom'].set_color('darkblue')
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        addition = ''
        if with_outliers:
            addition = ' without outliers'
        plt.title(f'{self.name}{addition}')
        plt.xlabel('Log10[Dilution]')

        plt.ylabel('OD450 nm')
        legend = plt.legend()
        legend.get_frame().set_facecolor('none')
        legend.get_frame().set_linewidth(0.0)

        plt.show()

        if folder_name is not None:
            plt.savefig(f'{folder_name}/{self.name} endpoint titer{addition}.png')

    def remove_sample(self, sample: Sample):
        assert sample.group == self
        self.samples.remove(sample)
        sample.group = None

    def __repr__(self):
        return f'AnalyticalGroup "{self.name}"\n' + \
               f'cutoff: {self.cutoff}\n' + \
               f'samples: {len(self.samples)}\n' + \
               '\n'.join([str(sample) for sample in self.samples]) + \
               f'average endpoint titer: {self.average_titer}\n' + \
               '\n--------'


def get_entry(data: pd.DataFrame, index: str) -> float:
    letter_index = index[0]
    number_index = int(index[1:])

    return data[number_index][letter_index]


# get set of the dilutions
def get_dilutions():
    # set dilutions
    dilutions = []
    user_set = input("Provide your dilution set: ")
    dilution_set = user_set.split(' ')

    # если введено первое разведение (1), а потом множитель (2) и число ячеек (3),
    # то можно сгенерировать разведение самим
    if len(dilution_set) == 3 and int(dilution_set[2]) <= 8:
        pre_dilutions = []
        current_group_indices = 1
        a = float(dilution_set[0])
        dilutions.append(math.log10(a))
        while current_group_indices < int(dilution_set[2]):
            a = a * float(dilution_set[1])
            dilutions.append(math.log10(a))
            current_group_indices += 1
    elif 3 < len(dilution_set) <= 8:
        # вводить числа, разделяя пробелами
        dilutions = [math.log10(float(x)) for x in dilution_set.split(' ')]
    else:
        print('Something wrong with your dilutions set')
        sys.exit()

    return dilutions


# find outliers in dataset
def detect_outlier(data: list, q1: int, q2: int):
    # find q1 and q3 values
    q1, q3 = np.percentile(sorted(data), [q1, q2])
    # compute IRQ
    iqr = q3 - q1
    # find lower and upper bounds
    lower_bound = q1 - (1.5 * iqr)
    upper_bound = q3 + (1.5 * iqr)

    outliers = [x for x in data if x <= lower_bound or x >= upper_bound]

    return outliers


# build test graph for debugging
def build_common_plot(groups: list, folder_name=None):
    data = []
    for group in groups:
        for sample in group.samples:
            if not sample.bad_data:
                a=[group.name, "%.0f" % sample.get_endpoint_titer(), sample.name]
                data.append(a)
    df = pd.DataFrame(np.array(data), columns=['Group', 'Titer', 'Sample'])
    df['Titer'] = df['Titer'].astype(np.int64)
    df_sorted = df.sort_values(['Titer', 'Group'])
    print(df_sorted)
    plt.figure(figsize=(8, 6), dpi=80)
    sns.stripplot(x='Group', y='Titer', data=df_sorted, color='black', size=5, jitter=0.1)

    for i in range(len(df['Group'].unique())-1):
        plt.vlines(i+.5, 0, 10, linestyles='solid', colors='black', alpha=0.2)

    plt.title('Average Endpoint Titer', fontsize=18)
    # plt.legend(title='Cylinders')
    plt.show()
    if folder_name is not None:
        plt.savefig(f"{folder_name}/Average Endpoint Titer.png")

# plate letters
letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

# get plate coordinates list
plate_coordinates_list = list()
for i in range(1, 13):
    for lit in letters:
        plate_coordinates_list.append(lit+str(i))


# write data to csv file
def write_data_to_csv(groups: list, folder_name=None, accuracy: float = cutoff_multiplier_accuracies[0]):
    with open(f'{folder_name}/data.csv', 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)

        for group in groups:
            writer.writerow(['Group'])
            writer.writerow([group.name])
            writer.writerow([' '])
            writer.writerow(['Cutoff'])
            writer.writerow([group.cutoff])
            writer.writerow([' '])
            writer.writerow(['Endpoint Titer'])
            print(f'{group.average_titer=}')
            writer.writerow([group.average_titer])
            writer.writerow([' '])
            writer.writerow(['Calculation Accuracy'])
            writer.writerow([f'{accuracy}'])
            writer.writerow(['Log10[Dil]'] + [s.name for s in group.samples if not s.bad_data])
            for i in range(0, len(group.samples[0].xdata)-1):
                writer.writerow([group.samples[0].xdata[i]] + [s.ydata[i] for s in group.samples if not s.bad_data])
            writer.writerow(['Endpoint titer'] + [s.endpoint_titer for s in group.samples if s not in group.outliers
                                                and not s.bad_data])
            writer.writerow(['*'*20])


def load_plate_data(file_path) -> pd.DataFrame:
    # get active sheet
    wb_obj = openpyxl.load_workbook(file_path)
    sheet = wb_obj.active

    # filter and transform data to dataframe
    rows = []
    for row in sheet.iter_rows(1, sheet.max_row):
        if row[0].value in letters:
            rows.append([x.value for x in row[1:]])
    return pd.DataFrame(rows, index=letters, columns=[x for x in range(1, 13)])


def main():
    # create folder for results storing
    current_directory = os.getcwd()
    final_directory_name = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    final_directory = os.path.join(current_directory, final_directory_name)
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)

    # get xlsx file with data and check if it is exist
    xlsx_path = sys.argv[1]
    if os.path.exists(xlsx_path):
        xlsx_file = Path(xlsx_path)
    else:
        print("Set xlsx file with data")
        return

    data = load_plate_data(xlsx_file)

    # вводить в формате А1-Н3 А4-Н8, разделяя пробелами
    groups_txt_input = [x.strip() for x in input("Set groups: ").split(' ')]

    # divide data into groups (we will get 2D array with coords)
    all_group_indices = list()
    for group in groups_txt_input:
        # divide user input by "-" to the start and end of the range
        start, end = group.split('-')
        # find the indexes of the start and end of the range and crop the array by these indexes
        all_group_indices.append(
            plate_coordinates_list[plate_coordinates_list.index(start):plate_coordinates_list.index(end) + 1])

    dilutions = get_dilutions()

    # create samples
    sample_groups = []
    # Словарь с группами, в качесве ключей и списком образцов, в качесвте значений
    samples_dict = {}
    for current_group_indices in all_group_indices:
        # use last entry index to determine which column letter is the last for each sample
        # (e.g. in A1-H3 it will be H, i.e. A1-H1, A2-H2, A3-H3)
        last_entry_index = current_group_indices[-1]
        last_column_letter = last_entry_index[0]

        samples = []
        current_sample_ydata = []
        for index in current_group_indices:
            current_sample_ydata.append(get_entry(data, index))
            # is this entry last in the current sample?
            if index.startswith(last_column_letter):
                # this entry was last, flush all accumulated ydata into new sample
                current_sample = Sample(f'sample {len(samples) + 1}', xdata=dilutions, ydata=current_sample_ydata)
                current_sample.get_popt_pcov()
                current_sample.get_R2()
                samples.append(current_sample)
                current_sample_ydata = []

        group = AnalyticalGroup(f'group {len(sample_groups) + 1}', samples)
        sample_groups.append(group)
        samples_dict[group] = [samples]

    for group in sample_groups:
        group.detect_outliers()
        group.plot_samples_data(folder_name=final_directory)
        if group.outliers:
            group.plot_samples_data(folder_name=final_directory, with_outliers=True)
        group.get_group_cutoff(99.0)
        group.calculate_average_titer()

    build_common_plot(sample_groups, folder_name=final_directory)
    write_data_to_csv(sample_groups, folder_name=final_directory)


if __name__ == '__main__':
    main()
